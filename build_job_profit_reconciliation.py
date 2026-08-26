import json, csv, re
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path('/home/ubuntu/sydneyautomationco-job-system')
SRC = ROOT / '.private-recovery' / 'data.baseline.json'
OUT = ROOT / 'reconciliation_output'
OUT.mkdir(exist_ok=True)

data = json.loads(SRC.read_text())
jobs = list(data.get('jobs', []) or [])
clients = {c.get('id'): c for c in data.get('clients', [])}
sup = list(data.get('supInvoices', []) or [])
expenses = data.get('expenses', []) or []

# User-confirmed overlay for the later Valdis job and invoice, absent from the
# protected baseline export. It is explicitly labelled so it cannot be mistaken
# for an original historical record.
valdis_client = next((c for c in clients.values() if c.get('company') == 'Valdis Berzins'), {})
if not any(j.get('jobNumber') == 'SAI_100097' for j in jobs):
    overlay_id = 'confirmed_sai_100097'
    jobs.append({'id': overlay_id, 'jobNumber': 'SAI_100097', 'type': 'SAI', 'status': 'paid',
                 'clientId': valdis_client.get('id'), 'quoteAmount': '2258.92',
                 'invoiceAmount': '2258.92', 'invoiceNumber': 'SAI_100097',
                 'gstExempt': False, 'paidDate': '2026-07-24',
                 'notes': 'User-confirmed overlay: after GST registration.'})
    sup.append({'id': 'confirmed_sinv_163', 'jobId': overlay_id, 'jobRef': 'SAI_100097',
                'supplier': 'J Lydement Electrical Pty Ltd', 'supplierName': 'J Lydement Electrical Pty Ltd',
                'ref': '163', 'date': '2026-07-24', 'amount': 935.88, 'costType': 'labour',
                'notes': 'User-confirmed invoice #163 for Valdis; GST-inclusive.'})
# Apply the user's confirmed Valdis mappings to existing baseline rows as an audit overlay.
valdis_allocations = {
    'SAI_100067': {'ref':'00000729', 'supplier':'Sharper Automation Pty Limited', 'amount':1788.46, 'gstExempt':True, 'costType':'parts'},
    'SAI_100070': {'ref':'102', 'supplier':'J Lydement Electrical Pty Ltd', 'amount':660.00, 'gstExempt':True, 'costType':'labour'},
    'SAI_100097': {'ref':'163', 'supplier':'J Lydement Electrical Pty Ltd', 'amount':935.88, 'gstExempt':False, 'costType':'labour'},
}
for job_number, alloc in valdis_allocations.items():
    job = next((j for j in jobs if j.get('jobNumber') == job_number), None)
    if not job: continue
    job['gstExempt'] = alloc['gstExempt']
    digits = ''.join(ch for ch in alloc['ref'] if ch.isdigit())
    inv = next((i for i in sup if i.get('jobId') == job.get('id') and ''.join(ch for ch in str(i.get('ref') or '') if ch.isdigit()) == digits), None)
    if not inv:
        inv = next((i for i in sup if i.get('jobId') == job.get('id') and not i.get('ref')), None)
    if inv:
        inv['ref'] = alloc['ref']; inv['amount'] = alloc['amount']; inv['supplier'] = alloc['supplier']; inv['supplierName'] = alloc['supplier']; inv['costType'] = alloc['costType']; inv['jobRef'] = job_number

timesheets = data.get('timesheets', []) or []
cfg = data.get('cfg', {}) or {}

def n(v):
    try: return float(v or 0)
    except (TypeError, ValueError): return 0.0

def money(v): return round(n(v), 2)

def client_name(j):
    c = clients.get(j.get('clientId'), {})
    return c.get('company') or c.get('name') or j.get('clientName') or j.get('clientId') or 'Unknown'

def revenue(j):
    if j.get('type') == 'SAQ' or str(j.get('jobNumber','')).startswith('SAQ') or j.get('status') != 'paid':
        return 0.0
    dep = max(0.0, n(j.get('depositAmount')))
    stored = n(j.get('invoiceAmount') or j.get('quoteAmount'))
    if not stored and not dep: return 0.0
    if dep > 0:
        inv = max(0.0, n(j.get('quoteAmount')) - dep)
    else:
        inv = stored if j.get('gstExempt') else (stored * 1.1 if n(j.get('quoteAmount')) > 0 and abs(n(j.get('quoteAmount'))-stored) > .02 and cfg.get('gstRegistered', True) is not False else stored)
    return money(dep + max(0.0, inv)) if dep else money(inv)

def base_date(j): return j.get('paidDate') or j.get('invoiceDate') or j.get('createdAt') or ''

def linked_supplier_rows(j):
    jid, ref = j.get('id'), j.get('jobNumber')
    return [i for i in sup if (jid and i.get('jobId') == jid) or (ref and i.get('jobRef') == ref)]

def linked_expense_rows(j):
    jid, ref = j.get('id'), j.get('jobNumber')
    return [i for i in expenses if (jid and i.get('jobId') == jid) or (ref and i.get('jobRef') == ref)]

def linked_labour_rows(j):
    jid, ref = j.get('id'), j.get('jobNumber')
    return [i for i in timesheets if (jid and i.get('jobId') == jid) or (ref and i.get('jobRef') == ref)]

# Only real service/invoice jobs are reconciled; quotes are kept in exceptions but not profit.
rows = []
for j in jobs:
    if j.get('type') == 'SAQ' or str(j.get('jobNumber','')).startswith('SAQ'):
        continue
    srows = linked_supplier_rows(j)
    erows = linked_expense_rows(j)
    lrows = linked_labour_rows(j)
    supplier_cost = money(sum(n(i.get('amount')) for i in srows))
    expense_cost = money(sum(n(i.get('amount')) for i in erows))
    labour_cost = money(sum(n(i.get('amount')) for i in lrows))
    actual_parts = money(j.get('actualPartsCost'))
    if actual_parts > 0 and supplier_cost > 0 and abs(actual_parts - supplier_cost) <= 0.05:
        # The supplier invoice is the source document for the same parts cost;
        # count it once and retain the stored actualPartsCost as corroboration.
        direct_parts = supplier_cost
        cost_basis = 'linked supplier invoice matches actualPartsCost; counted once'
        overlap_flag = False
    elif actual_parts > 0 and supplier_cost > 0:
        direct_parts = actual_parts
        cost_basis = 'actualPartsCost; supplier invoices also linked (review for overlap)'
        overlap_flag = True
    elif actual_parts > 0:
        direct_parts = actual_parts
        cost_basis = 'actualPartsCost'
        overlap_flag = False
    elif supplier_cost > 0:
        direct_parts = supplier_cost
        cost_basis = 'linked supplier invoices'
        overlap_flag = False
    else:
        direct_parts = 0.0
        cost_basis = 'none recorded'
        overlap_flag = False
    rev = revenue(j)
    direct = money(direct_parts + expense_cost + labour_cost)
    # Allocate only recurring overheads to paid jobs; this is management reporting, not statutory accounting.
    overhead_monthly = money(sum(n(x.get('monthly')) for x in cfg.get('overheads', []) if isinstance(x, dict)))
    overhead_alloc = money(overhead_monthly / max(1, sum(1 for x in jobs if x.get('status')=='paid' and base_date(x).startswith(base_date(j)[:7])))) if j.get('status')=='paid' and base_date(j)[:7] else 0.0
    gross_profit = money(rev - direct)
    operating_profit = money(gross_profit - overhead_alloc)
    flags=[]
    if j.get('status') == 'paid' and rev <= 0: flags.append('paid job has zero recognised revenue')
    if j.get('status') == 'paid' and direct <= 0: flags.append('no direct cost recorded')
    if j.get('status') in ('paid','invoiced','overdue') and not j.get('invoiceNumber'): flags.append('missing invoice reference')
    if overlap_flag: flags.append('actual parts cost and supplier invoices both present — check overlap')
    if j.get('depositAmount') and j.get('invoiceAmount') and n(j.get('quoteAmount')) > 0:
        if abs((n(j.get('depositAmount')) + n(j.get('invoiceAmount'))) - n(j.get('quoteAmount'))) > 0.02:
            flags.append('deposit plus balance does not equal quote')
    rows.append({
        'job_number': j.get('jobNumber',''), 'client': client_name(j), 'status': j.get('status',''), 'date': base_date(j),
        'invoice_number': j.get('invoiceNumber',''), 'revenue_gst_inc': rev, 'parts_supplier_cost': direct_parts,
        'other_expenses': expense_cost, 'labour_cost_recorded': labour_cost, 'direct_cost_total': direct,
        'allocated_overhead': overhead_alloc, 'gross_profit_before_overhead': gross_profit,
        'operating_profit_after_overhead': operating_profit,
        'gross_margin_pct': round(gross_profit/rev*100,1) if rev else '',
        'supplier_refs': '; '.join(str(i.get('ref') or '') for i in srows), 'expense_refs': '; '.join(str(i.get('id') or '') for i in erows),
        'cost_basis': cost_basis, 'flags': '; '.join(flags) if flags else 'OK'
    })

# Unlinked supplier/expense rows are leakage risks.
job_by_id = {j.get('id'): j for j in jobs}
job_by_ref = {j.get('jobNumber'): j for j in jobs}
unlinked_sup=[]
for i in sup:
    if not ((i.get('jobId') and i.get('jobId') in job_by_id) or (i.get('jobRef') and i.get('jobRef') in job_by_ref)):
        unlinked_sup.append(i)
unlinked_exp=[]
for i in expenses:
    if not ((i.get('jobId') and i.get('jobId') in job_by_id) or (i.get('jobRef') and i.get('jobRef') in job_by_ref)):
        unlinked_exp.append(i)

# User-confirmed current commitments, not present in the protected export; kept separate and provisional.
commitments = [
    ('JL Electrical — labour/electrician', 1799.00),
    ('Additional electrician labour — label to confirm', 1645.00),
    ('Sharper — parts supplier', 12912.03),
    ('Other parts/supplier amount — label to confirm', 1063.92),
]
commit_total = money(sum(v for _,v in commitments))

# Duplicate invoice references.
refs = [j.get('invoiceNumber') for j in jobs if j.get('invoiceNumber')]
dup_refs = sorted([r for r,c in Counter(refs).items() if c>1])

csv_path = OUT / 'job_profit_reconciliation.csv'
fieldnames = list(rows[0].keys()) if rows else []
with csv_path.open('w', newline='', encoding='utf-8') as f:
    w=csv.DictWriter(f, fieldnames=fieldnames); w.writeheader(); w.writerows(rows)

wb=Workbook(); ws=wb.active; ws.title='Job Reconciliation'
headers=fieldnames
ws.append(headers)
for r in rows: ws.append([r.get(h,'') for h in headers])
for cell in ws[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='1F4E78'); cell.alignment=Alignment(wrap_text=True)
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
for col in range(1, ws.max_column+1): ws.column_dimensions[get_column_letter(col)].width=min(42,max(12,len(str(ws.cell(1,col).value))+2))
for col_name in ['revenue_gst_inc','parts_supplier_cost','other_expenses','labour_cost_recorded','direct_cost_total','allocated_overhead','gross_profit_before_overhead','operating_profit_after_overhead']:
    if col_name in headers:
        c=headers.index(col_name)+1
        for cell in ws.iter_cols(min_col=c,max_col=c,min_row=2):
            for x in cell: x.number_format='$#,##0.00;[Red]-$#,##0.00'
if 'flags' in headers:
    c=headers.index('flags')+1
    for row in range(2,ws.max_row+1):
        if ws.cell(row,c).value != 'OK': ws.cell(row,c).fill=PatternFill('solid',fgColor='FFF2CC')

ws2=wb.create_sheet('Unallocated Items'); ws2.append(['Category','Description','Amount','Status'])
for i in unlinked_sup: ws2.append(['Supplier invoice', f"{i.get('supplier') or i.get('supplierName','')} / {i.get('ref','')}", n(i.get('amount')), 'No matching jobId/jobRef'])
for i in unlinked_exp: ws2.append(['Expense', f"{i.get('description','')} / {i.get('id','')}", n(i.get('amount')), 'No matching jobId/jobRef'])
for desc,amt in commitments: ws2.append(['User-confirmed commitment',desc,amt,'Not in protected export; provisional'])
for cell in ws2[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='1F4E78')
for row in ws2.iter_rows(min_row=2,min_col=3,max_col=3):
    for c in row: c.number_format='$#,##0.00;[Red]-$#,##0.00'
ws2.freeze_panes='A2'
for col in range(1,5): ws2.column_dimensions[get_column_letter(col)].width=[26,55,16,38][col-1]

summary = {
 'source': str(SRC), 'snapshot_savedAt': data.get('savedAt'), 'jobs_reconciled': len(rows),
 'paid_revenue': money(sum(r['revenue_gst_inc'] for r in rows)),
 'recorded_direct_costs': money(sum(r['direct_cost_total'] for r in rows)),
 'recorded_gross_profit': money(sum(r['gross_profit_before_overhead'] for r in rows)),
 'flagged_jobs': sum(1 for r in rows if r['flags']!='OK'), 'unlinked_supplier_count': len(unlinked_sup), 'unlinked_expense_count': len(unlinked_exp),
 'duplicate_invoice_refs': dup_refs, 'user_commitments_total': commit_total
}
wb.save(OUT / 'job_profit_reconciliation.xlsx')

# Management report.
report=[]
report += ['# Job-Level Profit Reconciliation', '', f"**Source:** protected baseline export `{SRC.name}` (snapshot `savedAt={data.get('savedAt')}`), plus the user-confirmed current commitments supplied in this conversation. This report does not alter source data.", '']
report += ['## Executive conclusion', '', f"The export contains **{len(rows)} non-quote jobs** available for reconciliation. Recorded paid revenue is **${summary['paid_revenue']:,.2f}**, against **${summary['recorded_direct_costs']:,.2f}** of matched direct costs, producing recorded gross profit of **${summary['recorded_gross_profit']:,.2f}** before allocated overhead. However, **{summary['flagged_jobs']} jobs are flagged** because costs are missing, deposit/balance treatment needs review, or cost records may overlap. Therefore the current data supports a positive recorded margin, but it is not yet a complete job-level profit statement until the flagged costs are resolved.", '']
report += ['## Reconciliation rules', '', 'Revenue follows the active application logic: paid service/invoice jobs only, GST-inclusive customer receipts, with deposit and balance components treated as one economic job where the data supports it. Supplier invoices are matched by `jobId` or `jobRef`; expenses are matched by `jobId` or `jobRef`; labour is matched from timesheets. Timesheets are empty in the protected export, so recorded labour cost is zero unless supplied through another source. A recurring overhead allocation is shown for management visibility only and is not a statutory accounting allocation.', '']
report += ['## Portfolio totals', '', '| Measure | Amount |', '|---|---:|', f"| Recorded paid revenue | ${summary['paid_revenue']:,.2f} |", f"| Matched direct costs | ${summary['recorded_direct_costs']:,.2f} |", f"| Recorded gross profit before overhead | ${summary['recorded_gross_profit']:,.2f} |", f"| User-confirmed current commitments not in export | ${commit_total:,.2f} |", f"| Unlinked supplier invoices | {len(unlinked_sup)} |", f"| Unlinked expenses | {len(unlinked_exp)} |", f"| Duplicate invoice references | {', '.join(dup_refs) if dup_refs else 'None'} |", '']
report += ['## Current commitments supplied separately', '', '| Description | Amount | Treatment |', '|---|---:|---|']
for desc,amt in commitments: report.append(f'| {desc} | ${amt:,.2f} | Provisional; not matched to a job in the protected export |')
report += [f'| **Total** | **${commit_total:,.2f}** | **Requires job matching before final profit conclusion** |', '']
report += ['## Highest-priority exceptions', '']
flagged=[r for r in rows if r['flags']!='OK']
for r in flagged[:30]: report.append(f"- **{r['job_number']} — {r['client']}**: {r['flags']}. Revenue ${r['revenue_gst_inc']:,.2f}; direct cost ${r['direct_cost_total']:,.2f}.")
if not flagged: report.append('No job-level exceptions were detected.')
report += ['', '## Interpretation', '', 'The key distinction is between **recorded profitability** and **cash safety**. A job can show positive gross profit while the business still feels short of cash if customers have not paid, suppliers have been paid early, or labour has not yet been entered. The four commitments supplied separately total $17,419.95 and should be matched to the jobs they relate to before relying on the displayed margins or drawing money from the operating account.', '']
report += ['## Required cleanup sequence', '', '1. Match each of the four current commitments to a job number and mark whether it has already been paid or remains payable. 2. Enter electrician/subcontractor labour against the relevant job rather than leaving labour at zero. 3. Review every flagged job with no direct cost and confirm whether it is genuinely labour-only or missing cost data. 4. Review any job where actual parts cost and supplier invoices are both present to prevent double counting. 5. Resolve duplicate invoice reference SAI_100041 only after confirming which invoice should retain it. 6. Re-run this reconciliation after the missing mappings are entered.', '']
report += ['## Files produced', '', '- `job_profit_reconciliation.xlsx` — workbook with job-level reconciliation and unallocated items.', '- `job_profit_reconciliation.csv` — flat job-level data for review or import.', '- `job_profit_reconciliation.md` — this management report.', '']
(OUT/'job_profit_reconciliation.md').write_text('\n'.join(report), encoding='utf-8')
print(json.dumps(summary, indent=2, ensure_ascii=False))
print('OUTPUT_DIR', OUT)
